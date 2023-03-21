from typing import Union, List

import openpyxl.utils.cell
from django.db.models import QuerySet, Prefetch
from django.shortcuts import get_object_or_404
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.survey_packages.models import (
    SurveyPackage,
    PackagePart,
    PackageSubjectSurvey,
    Respondent,
    PackageSubject,
)
from apps.survey_packages.serializers import (
    PackageContactSerializer,
    PackagePartSerializer,
    PackageSubjectSerializer,
    PackageSubjectSurveySerializer,
)
from apps.surveys.models import QuestionAnswer, SurveySector, SectorQuestion
from apps.workspaces.models import Workspace, RoutineDetail
from config.exceptions import InvalidInputException, InternalServerError


class SurveyPackageService(object):
    def __init__(self, package: Union[SurveyPackage, int]):
        if type(package) == int:
            self.package = get_object_or_404(SurveyPackage, id=package)
        else:
            self.package = package

    def add_contacts(self, contacts: List[dict]) -> SurveyPackage:
        if type(contacts) != list:
            raise InvalidInputException("'contacts' field must be a list")

        for c in contacts:
            serializer = PackageContactSerializer(data=c)
            if serializer.is_valid(raise_exception=True):
                serializer.save(survey_package_id=self.package.id)

        return self.package

    def delete_related_components(self) -> None:
        PackagePart.objects.filter(survey_package_id=self.package.id).delete()

    def create_parts(self, data: list[dict], package_id: int) -> list[PackagePart]:
        created_parts = []
        for d in data:
            part = self.create_part(d, package_id=package_id)
            created_parts.append(part)

        return created_parts

    def create_part(self, part_data: dict, package_id: int) -> PackagePart:
        # Create a part
        if "title" not in part_data or part_data["title"] is None:
            raise InvalidInputException("a part should have a title")

        serializer = PackagePartSerializer(data=dict(title=part_data["title"]))
        if serializer.is_valid(raise_exception=True):
            part = serializer.save(survey_package_id=package_id)

        # Create subjects
        if "subjects" in part_data or part_data["subjects"] is not None:
            self._create_subjects(part_data["subjects"], part.id)
            part.refresh_from_db()

        return part

    def _create_subjects(self, subjects_list: list[dict], part_id: int):
        for subject in subjects_list:
            serializer = PackageSubjectSerializer(data=subject)
            if serializer.is_valid(raise_exception=True):
                serializer.save(package_part_id=part_id)

    @staticmethod
    def associate_subject_with_surveys(subject_id: int, data: list[dict]) -> None:
        for ss in data:
            serializer = PackageSubjectSurveySerializer(data=ss)
            if serializer.is_valid(raise_exception=True):
                serializer.save(subject_id=subject_id, survey_id=ss.get("survey", None))

    @staticmethod
    def delete_related_surveys(subject_id: int) -> None:
        PackageSubjectSurvey.objects.filter(subject_id=subject_id).delete()


class ResponseExportService(object):
    def __init__(self, workspace: Workspace, survey_package: SurveyPackage):
        self.workspace = workspace
        self.survey_package = survey_package
        self.respondents = None
        self.workbook = None
        self.worksheet = None
        self._col_num = 6

    def _base_data(self) -> list[Union[str, int]]:
        routine_id: int = self.workspace.routine.id
        routine_detail: RoutineDetail = RoutineDetail.objects.filter(
            routine_id=routine_id, survey_package_id=self.survey_package.id
        ).first()

        workspace_name: str = self.workspace.name
        package_name = self.survey_package.title
        nth_day = routine_detail.nth_day
        time = routine_detail.time

        return [workspace_name, package_name, nth_day, time]

    def _create_worksheet_template(self) -> Worksheet:
        wb: Workbook = Workbook()
        self.workbook = wb
        ws: Worksheet = wb.active
        self.worksheet = ws

        self.worksheet.append(["워크스페이스", "설문 제목", "차시 (일)", "응답 지정 일시", "피험자ID"])
        self.worksheet.append(self._base_data())
        respondent_list = self._set_respondents_list()

        i = 0
        for row in self.worksheet.iter_rows(
            min_row=2, min_col=5, max_col=5, max_row=len(respondent_list) + 1
        ):
            for cell in row:
                cell.value = respondent_list[i]
            i += 1

        return self.worksheet

    def _set_respondents_list(self) -> list[dict]:
        respondents: list[dict] = (
            Respondent.objects.filter(
                survey_package_id=self.survey_package.id, workspace_id=self.workspace.id
            )
            .order_by("respondent_id")
            .all()
            .values_list("respondent_id", flat=True)
        )
        self.respondents = respondents

        return self.respondents

    def _add_sector_data(self, sector_data, prefix):
        question_type = sector_data.question_type
        questions = sector_data.questions.all()

        for question in questions:
            col_letter = openpyxl.utils.cell.get_column_letter(self._col_num)

            # bread crumb format question number
            if self.worksheet[f"{col_letter}1"].value is None:
                formatted_question_number = str(question.number)

                if formatted_question_number[-1] != "0":
                    formatted_question_number = formatted_question_number.replace(
                        ".", "-"
                    )
                else:
                    dot_index = formatted_question_number.index(".")
                    formatted_question_number = formatted_question_number[:dot_index]

                self.worksheet[
                    f"{col_letter}1"
                ] = f"{prefix}-{formatted_question_number}"

            print(self.respondents, question.answers.count())
            if len(self.respondents) != question.answers.count():
                raise InternalServerError()

            row = 2
            for answer in question.answers.all():
                answer_val = answer.answer
                if "$" in answer_val:
                    answer_val = answer_val.replace("$", ", ")

                if question_type in ["likert", "extent", "single_select"]:
                    try:
                        answer_val = int(answer_val)
                        self.worksheet[f"{col_letter}{row}"].number_format = "0"
                        self.worksheet[f"{col_letter}{row}"].value = answer_val
                    except Exception:
                        pass

                self.worksheet[f"{col_letter}{row}"] = answer_val

                row += 1

            self._col_num += 1

    def _add_survey_data(self, survey_data, prefix):
        if survey_data.number and survey_data.number != "":
            prefix = f"{prefix}-{survey_data.number}"

        survey = survey_data.survey
        prefix = f"{prefix}-{survey.abbr}"

        sectors = survey.sectors.all()

        for sector in sectors:
            self._add_sector_data(sector, prefix)

    def _add_subject_data(self, subject_data):
        prefix = f"{subject_data.number}"
        for survey in subject_data.surveys.all():
            self._add_survey_data(survey, prefix)

    def _add_part_data(self, part_data: PackagePart):
        print(part_data.subjects)
        for subject in part_data.subjects.all():
            self._add_subject_data(subject)

    def export_to_worksheet(self):
        self._create_worksheet_template()

        queryset: QuerySet = PackagePart.objects.filter(
            survey_package_id=self.survey_package.id
        ).prefetch_related(
            Prefetch(
                "subjects",
                queryset=PackageSubject.objects.prefetch_related(
                    Prefetch(
                        "surveys",
                        queryset=PackageSubjectSurvey.objects.select_related(
                            "survey"
                        ).prefetch_related(
                            Prefetch(
                                "survey__sectors",
                                queryset=SurveySector.objects.prefetch_related(
                                    Prefetch(
                                        "questions",
                                        queryset=SectorQuestion.objects.prefetch_related(
                                            Prefetch(
                                                "answers",
                                                queryset=QuestionAnswer.objects.filter(
                                                    survey_package_id=self.survey_package.id
                                                ).order_by("respondent_id"),
                                            )
                                        ).order_by(
                                            "number"
                                        ),
                                    )
                                ),
                            )
                        ),
                    )
                ),
            )
        )

        for part in queryset:
            self._add_part_data(part)

        return self.workbook


class SurveyPackageExportService(object):
    def __init__(self, survey_package_id: int):
        self.survey_package_id = survey_package_id
        self.workbook = None
        self.worksheet = None

    def _get_queryset(self):
        return PackagePart.objects.filter(
            survey_package_id=self.survey_package_id
        ).prefetch_related(
            Prefetch(
                "subjects",
                queryset=PackageSubject.objects.prefetch_related(
                    Prefetch(
                        "surveys",
                        queryset=PackageSubjectSurvey.objects.select_related(
                            "survey"
                        ).prefetch_related(
                            Prefetch(
                                "survey__sectors",
                                queryset=SurveySector.objects.prefetch_related(
                                    "common_choices",
                                    Prefetch(
                                        "questions",
                                        queryset=SectorQuestion.objects.prefetch_related(
                                            "choices"
                                        ),
                                    ),
                                ),
                            )
                        ),
                    )
                ),
            )
        )

    def _create_worksheet_template(self) -> None:
        wb: Workbook = Workbook()
        self.workbook = wb
        ws: Worksheet = wb.active
        self.worksheet = ws

        self.worksheet.append(
            ["구분", "대주제", "소주제", "문제유형", "연결섹터여부", "공통선지", "문항번호", "문항내용", "문항선지"]
        )

    def _add_row(self, row_data: dict):
        self.worksheet.append(list(row_data.values()))
        print("appended row")

    def _add_question(self, question_data: SectorQuestion, row_data: dict):
        row_data = dict(
            **row_data,
            question_number=question_data.number,
            question_content=question_data.content,
        )

        question_choices = question_data.choices.all()

        question_choice_str = ""
        if question_choices is not None or question_data.choices.count() != 0:
            for choice in question_data.choices.all():
                content = ""
                if choice.content is not None:
                    content += choice.content
                if choice.is_descriptive:
                    content += choice.desc_form

                content = content.replace("%d", "[숫자]")
                content = content.replace("%s", "[문자]")

                question_choice_str += f"{choice.number}. {content}/"
            question_choice_str = question_choice_str[:-1]

        row_data["question_choices"] = question_choice_str
        self._add_row(row_data)

    def _add_sector(self, sector: SurveySector, row_data: dict):
        row_data = dict(**row_data, question_type=sector.question_type)
        if sector.is_linked is True:
            row_data["is_linked"] = "Y"
        else:
            row_data["is_linked"] = "N"

        common_choices_str = ""
        if sector.common_choices is not None:
            for common_choice in sector.common_choices.all():
                common_choices_str += (
                    f"{common_choice.number}. {common_choice.content}/"
                )
            common_choices_str = common_choices_str[:-1]

        row_data["common_choices"] = common_choices_str

        for question in sector.questions.all():
            self._add_question(question, row_data)

    def _add_subject_survey(self, subject_survey: PackageSubjectSurvey, row_data: dict):
        if subject_survey.title is not None:
            row_data = dict(**row_data, subject_survey_title=subject_survey.title)
        else:
            row_data = dict(**row_data, subject_survey_title="")

        survey = subject_survey.survey

        for sector in survey.sectors.all():
            self._add_sector(sector, row_data)

    def _add_subject(self, subject: PackageSubject, part_title: str):
        row_data = dict(part_title=part_title, subject_title=subject.title)
        for subject_survey in subject.surveys.all():
            self._add_subject_survey(subject_survey, row_data)

    def _add_part(self, part_data: PackagePart):
        title = part_data.title
        for subject in part_data.subjects.all():
            self._add_subject(subject, title)

    def export_to_workbook(self):
        self._create_worksheet_template()
        queryset = self._get_queryset()

        for part in queryset:
            self._add_part(part)

        return self.workbook
